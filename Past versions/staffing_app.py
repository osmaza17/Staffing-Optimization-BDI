import flet as ft
from pulp import *
import json
import os
import threading
import time
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import platform
import subprocess

# File to save data
DATA_FILE = "staffing_data.json"

# =============================================================================
# DATA AND MODEL FUNCTIONS
# =============================================================================

def load_data():
    """Loads data from the JSON file if it exists."""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None

def save_data(data):
    """Saves data to the JSON file."""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def solve_model(data):
    """
    Builds and solves the optimization model using PuLP.
    """
    people = data['people']
    tasks = data['tasks']
    hours = data['hours']
    D = data['D']  # Availability
    Q = data['Q']  # Qualifications/Skills
    R = data['R']  # Requirements
    F = data['F']  # Fixed/Mandatory assignments
    
    # Parameters
    alpha = float(data['alpha'])
    beta = float(data['beta'])
    gamma = float(data['gamma'])
    epsilon = float(data['epsilon'])
    timelimit = int(data['timelimit'])
    
    print("Creating model...")
    model = LpProblem("Staffing", LpMinimize)
    
    # Variables
    X = LpVariable.dicts("X", (people, tasks, hours), cat='Binary')
    W = LpVariable.dicts("W", people, lowBound=0, cat='Integer')
    W_max = LpVariable("W_max", lowBound=0)
    W_min = LpVariable("W_min", lowBound=0)
    
    hours_minus_last = hours[:-1]
    Y = LpVariable.dicts("Y", (people, tasks, hours_minus_last), cat='Binary')
    
    hours_minus_first = hours[1:]
    S = LpVariable.dicts("S", (people, hours_minus_first), cat='Binary')
    
    U = LpVariable.dicts("U", (people, tasks, hours), cat='Binary')
    
    # Objective Function
    model += (
        alpha * (W_max - W_min) +
        beta * lpSum(Y[i][t][h] for i in people for t in tasks for h in hours_minus_last) +
        gamma * lpSum(S[i][h] for i in people for h in hours_minus_first) +
        epsilon * lpSum(U[i][t][h] for i in people for t in tasks for h in hours)
    )
    
    # Constraints
    for i in people:
        for h in hours:
            model += lpSum(X[i][t][h] for t in tasks) <= D[i][h]
    
    for t in tasks:
        for h in hours:
            model += lpSum(X[i][t][h] for i in people) == R[t][h]
    
    for i in people:
        for t in tasks:
            for h in hours:
                model += X[i][t][h] <= Q[i][t]
    
    for i in people:
        model += W[i] == lpSum(X[i][t][h] for t in tasks for h in hours)
        model += W_max >= W[i]
        model += W_min <= W[i]
    
    for i in people:
        for t in tasks:
            for h in hours_minus_last:
                h_next = hours[hours.index(h) + 1]
                model += Y[i][t][h] >= X[i][t][h] + X[i][t][h_next] - 1
    
    for i in people:
        for h in hours_minus_first:
            h_prev = hours[hours.index(h) - 1]
            T_ih = lpSum(X[i][t][h] for t in tasks)
            T_ih_prev = lpSum(X[i][t][h_prev] for t in tasks)
            model += S[i][h] >= T_ih - T_ih_prev
    
    for i in people:
        for t in tasks:
            for h in hours:
                model += U[i][t][h] >= F[i][t][h] - X[i][t][h]
    
    print(f"Running CBC solver (time limit: {timelimit}s)...")
    model.solve(PULP_CBC_CMD(msg=1, timeLimit=timelimit))
    
    return model, X, W, W_max, W_min

# =============================================================================
# MAIN APP (FLET)
# =============================================================================

class StaffingApp:
    def __init__(self):
        self.data = load_data()
        self.possible_hours = [16, 17, 18, 19, 20, 21, 22, 23, 0, 1, 2, 3, 4, 5, 6, 7, 8]
        
        # --- COLORS ---
        self.COLOR_ACTIVE = "#C6EFCE"
        self.TEXT_ACTIVE = "#006100"      
        self.COLOR_INACTIVE = "#FFC7CE"
        self.TEXT_INACTIVE = "#9C0006"    
        self.COLOR_HEADER_BG = "#F2F2F2"
        self.COLOR_BORDER = "#bfbfbf"     
        self.COLOR_TEXT_HIGHLIGHT = "blue700" 
        self.COLOR_BG_PANEL = "white"        
        self.COLOR_NEUTRAL = "grey200"

        self.available_colors = ["blue200", "red200", "green200", "amber200", "purple200", "cyan200", "orange200", "pink200", "teal200", "indigo200", "lime200", "brown200"]
        self.task_colors = {}

        self.FLET_TO_HEX = {
            "blue200": "90CAF9", "red200": "EF9A9A", "green200": "A5D6A7", 
            "amber200": "FFE082", "purple200": "CE93D8", "cyan200": "80DEEA", 
            "orange200": "FFCC80", "pink200": "F48FB1", "teal200": "80CBC4", 
            "indigo200": "9FA8DA", "lime200": "E6EE9C", "brown200": "BCAAA4",
            "white": "FFFFFF", "red100": "FFCDD2" 
        }

        self.state_hours = {} 
        self.state_D = {}
        self.state_Q = {}
        self.state_R = {} 
        self.state_F = {}
        
        # Diccionario para navegación por teclado en la matriz R: (row_idx, col_idx) -> TextField
        self.r_cells = {}
        
        self.grid_controls = {'D': {}, 'Q': {}} 
        self.bulk_states = {} 

        self.input_people_val = ""
        self.input_tasks_val = ""
        
        for i in range(17):
            active = 0
            if self.data and 'hours' in self.data and i in self.data['hours']:
                active = 1
            self.state_hours[i] = active

        self.status_text = ft.Text("Ready.", color="grey600", size=12)
        self.progress_bar = ft.ProgressBar(width=200, color="blue", bgcolor="#eeeeee", visible=False)
        self.btn_optimize = None
        
        self.content_matrices = ft.Column(scroll=ft.ScrollMode.ALWAYS, expand=True)
        self.content_mandatory = ft.Column(scroll=ft.ScrollMode.ALWAYS, expand=True)
        self.container_hours = ft.Column()
        self.page = None

    def main(self, page: ft.Page):
        self.page = page
        self.page.title = "Staffing Optimizer - Excel Mode"
        self.page.window.width = 1050
        self.page.window.height = 700
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = 10

        def_pers = '\n'.join(self.data.get('people', [])) if self.data else ""
        def_task = '\n'.join(self.data.get('tasks', [])) if self.data else ""
        self.input_people_val = def_pers
        self.input_tasks_val = def_task

        self.txt_people = ft.TextField(
            value=def_pers, multiline=True, expand=True, text_size=12, border=ft.InputBorder.NONE,
            on_blur=self.on_input_change
        )
        self.txt_tasks = ft.TextField(
            value=def_task, multiline=True, expand=True, text_size=12, border=ft.InputBorder.NONE,
            on_blur=self.on_input_change
        )
        
        self.generate_hour_buttons()

        def get_p(k, d): return self.data.get(k, d) if self.data else d
        
        def input_param(label, value):
            return ft.Row([
                ft.Container(
                    content=ft.Text(label, size=11, color="grey800", weight="bold", text_align="right"),
                    width=140, alignment=ft.alignment.center_right, padding=ft.padding.only(right=10)
                ),
                ft.Container(
                    content=ft.TextField(
                        value=value, text_size=12, border=ft.InputBorder.NONE, height=30, 
                        content_padding=ft.padding.only(left=10, right=10, bottom=14), text_align=ft.TextAlign.LEFT
                    ),
                    border=ft.border.all(1, self.COLOR_BORDER), border_radius=5, bgcolor="white",
                    height=28, width=100, alignment=ft.alignment.center_left
                )
            ], spacing=0, vertical_alignment=ft.CrossAxisAlignment.CENTER)

        self.in_alpha = input_param("Alpha (Equity)", str(get_p('alpha', 1.0)))
        self.in_beta = input_param("Beta (Continuity)", str(get_p('beta', 0.3)))
        self.in_gamma = input_param("Gamma (Switches)", str(get_p('gamma', 0.5)))
        self.in_epsilon = input_param("Epsilon (Fixed)", str(get_p('epsilon', 10.0)))
        self.in_timelimit = input_param("Max Time (s)", str(get_p('timelimit', 60)))

        config_content = ft.Row(
            controls=[
                ft.Container(
                    expand=1,
                    content=ft.Column([
                        ft.Text("1. People List", color=self.COLOR_TEXT_HIGHLIGHT, weight="bold"),
                        ft.Container(
                            content=self.txt_people,
                            border=ft.border.all(1, self.COLOR_BORDER),
                            border_radius=5, bgcolor="white", padding=5, expand=True
                        )
                    ], expand=True)
                ),
                ft.VerticalDivider(width=1, color=self.COLOR_BORDER),
                ft.Container(
                    expand=1,
                    content=ft.Column([
                        ft.Text("2. Task List", color=self.COLOR_TEXT_HIGHLIGHT, weight="bold"),
                        ft.Container(
                            content=self.txt_tasks,
                            border=ft.border.all(1, self.COLOR_BORDER),
                            border_radius=5, bgcolor="white", padding=5, expand=True
                        )
                    ], expand=True)
                ),
                ft.VerticalDivider(width=1, color=self.COLOR_BORDER),
                ft.Container(
                    expand=1.5,
                    content=ft.Column([
                        ft.Text("3. Active Hours", color=self.COLOR_TEXT_HIGHLIGHT, weight="bold"),
                        ft.Container(
                            content=self.container_hours, 
                            border=ft.border.all(1, self.COLOR_BORDER),
                            padding=10, border_radius=5, bgcolor="white"
                        ),
                        ft.Divider(),
                        ft.Text("4. Parameters", color=self.COLOR_TEXT_HIGHLIGHT, weight="bold"),
                        ft.Column([self.in_alpha, self.in_beta, self.in_gamma, self.in_epsilon, self.in_timelimit], spacing=5)
                    ], spacing=10, scroll=ft.ScrollMode.AUTO)
                )
            ],
            expand=True,
            vertical_alignment=ft.CrossAxisAlignment.START,
            spacing=20
        )

        self.tabs_control = ft.Tabs(
            selected_index=0, animation_duration=0, expand=True,
            tabs=[
                ft.Tab(text="1. Configuration", content=ft.Container(content=config_content, padding=10)),
                ft.Tab(text="2. Data Matrices", content=ft.Container(content=self.content_matrices, padding=10)),
                ft.Tab(text="3. Mandatory/Fixed", content=ft.Container(content=self.content_mandatory, padding=10)),
            ]
        )

        self.btn_optimize = ft.ElevatedButton(
            "CALCULATE (OPTIMIZE)", on_click=self.run_optimization_thread, 
            height=40, width=200, style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=2), color="white", bgcolor="blue")
        )

        self.page.add(
            ft.Container(content=self.tabs_control, expand=True),
            ft.Container(
                content=ft.Row([
                    ft.Row([
                        self.btn_optimize,
                        ft.VerticalDivider(width=20),
                        self.progress_bar, 
                        self.status_text
                    ], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Row([
                        ft.Text("Developed by ", size=10, color="grey400", italic=True),
                        ft.TextButton(
                            content=ft.Text("Óscar Martínez Zamora", size=10, color="blue400", italic=True),
                            url="https://www.linkedin.com/in/oscarmartinezzamora/",
                            style=ft.ButtonStyle(padding=0, overlay_color="transparent"),
                            tooltip="Ver perfil en LinkedIn"
                        ),
                        ft.IconButton(
                            icon=ft.Icons.LINK,
                            icon_size=14,
                            icon_color="blue400",
                            url="https://www.linkedin.com/in/oscarmartinezzamora/",
                            tooltip="LinkedIn",
                            style=ft.ButtonStyle(padding=0),
                        )
                    ], spacing=0, vertical_alignment=ft.CrossAxisAlignment.CENTER)
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                padding=10, bgcolor=self.COLOR_BG_PANEL,
                border=ft.border.only(top=ft.border.BorderSide(1, self.COLOR_BORDER))
            )
        )

        self.generate_tables()

    def generate_hour_buttons(self):
        row1 = ft.Row(wrap=False, spacing=2) 
        row2 = ft.Row(wrap=False, spacing=2)
        split_idx = 9
        for i, h in enumerate(self.possible_hours):
            is_active = self.state_hours.get(i, 0)
            btn = ft.Container(
                content=ft.Text(f"{h:02d}h", size=11, weight=ft.FontWeight.BOLD, color="white" if is_active else "black"),
                width=38, height=25, bgcolor="#217346" if is_active else self.COLOR_NEUTRAL, 
                alignment=ft.alignment.center, border_radius=5, on_click=lambda e, idx=i: self.toggle_hour(e, idx)
            )
            if i < split_idx: row1.controls.append(btn)
            else: row2.controls.append(btn)
        self.container_hours.controls = [row1, row2]
        if self.page: self.page.update()

    def toggle_hour(self, e, idx):
        curr = self.state_hours[idx]
        self.state_hours[idx] = 1 - curr
        is_active = self.state_hours[idx]
        e.control.bgcolor = "#217346" if is_active else self.COLOR_NEUTRAL
        e.control.content.color = "white" if is_active else "black"
        e.control.update()
        self.generate_tables()

    def on_input_change(self, e):
        if e.control == self.txt_people:
            if self.input_people_val == self.txt_people.value: return
            self.input_people_val = self.txt_people.value
        elif e.control == self.txt_tasks:
            if self.input_tasks_val == self.txt_tasks.value: return
            self.input_tasks_val = self.txt_tasks.value
        self.generate_tables()

    def create_cell_button(self, label_active, label_inactive, tipo, k1, k2, k3=None, width=60, height=28):
        return self.create_cell_button_scaled(label_active, label_inactive, tipo, k1, k2, k3, width, height, font_size=10)

    def toggle_matrix_btn(self, e):
        data = e.control.data
        self._update_single_cell(e.control, data['tipo'], data['k1'], data['k2'], data['k3'])

    def _update_single_cell(self, control, tipo, k1, k2, k3, force_val=None):
        current_val = 0
        if force_val is not None:
            pass 
        else:
            if tipo == 'D': current_val = self.state_D[k1][k2]
            elif tipo == 'Q': current_val = self.state_Q[k1][k2]
            elif tipo == 'F': current_val = self.state_F[k1][k2][k3]
            force_val = 1 - current_val

        new_val = force_val
        if tipo == 'D': self.state_D[k1][k2] = new_val
        elif tipo == 'Q': self.state_Q[k1][k2] = new_val
        elif tipo == 'F': self.state_F[k1][k2][k3] = new_val
        
        bg_color = "white"
        text_color = "grey"
        text_value = ""

        if new_val:
            if tipo == 'F':
                bg_color = self.task_colors.get(k2, self.COLOR_ACTIVE)
                text_value = "" 
            else:
                bg_color = self.COLOR_ACTIVE
                text_color = self.TEXT_ACTIVE
                text_value = "YES"
        else:
            if tipo == 'D' or tipo == 'Q':
                bg_color = self.COLOR_INACTIVE
                text_color = self.TEXT_INACTIVE
                text_value = "NO"
            elif tipo == 'F':
                bg_color = "white"
                text_value = ""
        
        control.bgcolor = bg_color
        control.content.color = text_color
        control.content.value = text_value
        control.update()

    def _get_val_from_memory_or_json(self, tipo, k1, k2=None, k3=None):
        if tipo == 'D':
            if k1 in self.state_D and k2 in self.state_D[k1]: return self.state_D[k1][k2]
            if self.data and 'D' in self.data: return int(self.data['D'].get(k1, {}).get(str(k2), 1))
            return 1
        if tipo == 'Q':
            if k1 in self.state_Q and k2 in self.state_Q[k1]: return self.state_Q[k1][k2]
            if self.data and 'Q' in self.data: return int(self.data['Q'].get(k1, {}).get(k2, 1))
            return 1
        if tipo == 'F':
            if k1 in self.state_F and k2 in self.state_F[k1] and k3 in self.state_F[k1][k2]: return self.state_F[k1][k2][k3]
            if self.data and 'F' in self.data: return int(self.data['F'].get(k1, {}).get(k2, {}).get(str(k3), 0))
            return 0
        if tipo == 'R':
            if k1 in self.state_R and k2 in self.state_R[k1]: return self.state_R[k1][k2]
            # === CAMBIO: Valor por defecto de JSON o memoria si no existe es 0 (vacío) ===
            if self.data and 'R' in self.data: return int(self.data['R'].get(k1, {}).get(str(k2), 0))
            return 0
        return 0

    def create_bulk_action_cell(self, action_type, matrix_type, key, width=28, height=28):
        icon = ft.Icons.SWAP_HORIZ if action_type == 'row' else ft.Icons.SWAP_VERT
        tooltip = f"{action_type} Action {matrix_type}"
        
        if matrix_type == 'R':
            icon = ft.Icons.EXPOSURE_ZERO
            tooltip = "Reset to 0"

        return ft.Container(
            width=width, height=height, 
            content=ft.Icon(icon, size=16, color="grey"), 
            bgcolor="#eeeeee", border_radius=3, alignment=ft.alignment.center,
            tooltip=tooltip,
            on_click=lambda e: self.execute_bulk_action(action_type, matrix_type, key),
            ink=True
        )

    def execute_bulk_action(self, action_type, matrix_type, key):
        if matrix_type == 'R':
            cols = self.indices_horas
            rows = self.tasks
            
            # Reset values in R
            def reset_cell_R(t, h):
                # Update memory
                self.state_R.setdefault(t, {})[h] = 0
                
                # Update UI via the TextField if it exists
                try:
                    t_idx = self.tasks.index(t)
                    h_idx = self.indices_horas.index(h)
                    tf = self.r_cells.get((t_idx, h_idx))
                    if tf:
                        # === CAMBIO: Limpiar visualmente a "" ===
                        tf.value = ""
                        tf.update()
                except:
                    pass

            if action_type == 'row':
                for c in cols: reset_cell_R(key, c)
            elif action_type == 'col':
                for r in rows: reset_cell_R(r, key)
            return
        
        target_dict = None
        state_dict = None
        
        if matrix_type == 'D':
            cols = self.indices_horas
            rows = self.people
            target_dict = self.grid_controls['D']
            state_dict = self.state_D
        elif matrix_type == 'Q':
            cols = self.tasks
            rows = self.people
            target_dict = self.grid_controls['Q']
            state_dict = self.state_Q
        else:
            return
        
        current_val = 1 
        if action_type == 'row':
            first_col = cols[0] if cols else None
            if first_col is not None:
                current_val = state_dict.get(key, {}).get(first_col, 1)
        elif action_type == 'col':
            first_row = rows[0] if rows else None
            if first_row is not None:
                current_val = state_dict.get(first_row, {}).get(key, 1)
        
        new_val = 1 - current_val

        if action_type == 'row':
            for c in cols:
                ctrl = target_dict.get((key, c))
                if ctrl:
                    self._update_single_cell(ctrl, matrix_type, key, c, None, force_val=new_val)
        
        elif action_type == 'col':
            for r in rows:
                ctrl = target_dict.get((r, key))
                if ctrl:
                    self._update_single_cell(ctrl, matrix_type, r, key, None, force_val=new_val)

    def generate_tables(self):
        # 1. Update persisted R state from current R cells before destroy
        # This is redundant now because create_excel_input updates state on_change, 
        # but kept as safety.
        pass

        self.people = [t.strip() for t in self.txt_people.value.split('\n') if t.strip()]
        self.tasks = [t.strip() for t in self.txt_tasks.value.split('\n') if t.strip()]
        self.indices_horas = [i for i, v in self.state_hours.items() if v == 1]
        self.indices_horas.sort()

        self.task_colors = {}
        for i, t in enumerate(self.tasks):
            color = self.available_colors[i % len(self.available_colors)]
            self.task_colors[t] = color

        self.grid_controls = {'D': {}, 'Q': {}}
        self.r_cells = {} # Reset navigation map

        if not self.people or not self.tasks or not self.indices_horas:
            self.status_text.value = "Warning: Missing data (people, tasks or hours)."
            if self.page: self.status_text.update()
            return

        self.status_text.value = "Regenerating spreadsheet view..."
        if self.page: self.status_text.update()

        CELL_W_NAME = 100 
        CELL_W_HOUR = 60  
        CELL_W_TASK = 60 
        CELL_W_TASK_LABEL = 60 
        CELL_W_ACTION = 20 
        CELL_H = 20
        FONT_SIZE = 10
        CELL_W_R = 70 

        container_d = ft.Column(spacing=2)
        container_r = ft.Column(spacing=2)
        container_q = ft.Column(spacing=2)
        container_f = ft.Column(spacing=2)

        def cell_header(text, width):
            return ft.Container(
                width=width, height=CELL_H,
                content=ft.Text(text, size=FONT_SIZE, weight="bold", color="black"), 
                alignment=ft.alignment.center,
                bgcolor=self.COLOR_HEADER_BG,
                border=None, border_radius=3
            )

        def cell_name(text, width=CELL_W_NAME): 
            return ft.Container(
                width=width, height=CELL_H,
                content=ft.Text(text, size=FONT_SIZE, color="black"), 
                alignment=ft.alignment.center_left, padding=ft.padding.only(left=5),
                bgcolor="white", border=None, border_radius=3
            )

        # 1. AVAILABILITY (D)
        rows_d = []
        header_actions = [ft.Container(width=CELL_W_ACTION + CELL_W_NAME, height=CELL_H)]
        for h in self.indices_horas:
            header_actions.append(self.create_bulk_action_cell('col', 'D', h, width=CELL_W_HOUR, height=CELL_H))
        rows_d.append(ft.Row(controls=header_actions, spacing=2))

        header_controls = [ft.Container(width=CELL_W_ACTION, height=CELL_H), cell_header("Person", CELL_W_NAME)]
        for h in self.indices_horas:
            header_controls.append(cell_header(f"{self.possible_hours[h]:02d}h", CELL_W_HOUR))
        rows_d.append(ft.Row(controls=header_controls, spacing=2))

        for pers in self.people:
            row_ctrls = [
                self.create_bulk_action_cell('row', 'D', pers, width=CELL_W_ACTION, height=CELL_H),
                cell_name(pers)
            ]
            for h in self.indices_horas:
                row_ctrls.append(self.create_cell_button_scaled("YES", "NO", 'D', pers, h, width=CELL_W_HOUR, height=CELL_H, font_size=FONT_SIZE))
            rows_d.append(ft.Row(controls=row_ctrls, spacing=2))

        # 2. REQUIREMENTS (R) - EXCEL STYLE
        rows_r = []
        header_r_act = [ft.Container(width=CELL_W_ACTION + CELL_W_TASK_LABEL, height=CELL_H)]
        for h in self.indices_horas:
            header_r_act.append(self.create_bulk_action_cell('col', 'R', h, width=CELL_W_R, height=CELL_H))
        rows_r.append(ft.Row(controls=header_r_act, spacing=2))

        header_r = [ft.Container(width=CELL_W_ACTION, height=CELL_H), cell_header("Task", CELL_W_TASK_LABEL)]
        for h in self.indices_horas:
            header_r.append(cell_header(f"{self.possible_hours[h]:02d}h", CELL_W_R)) 
        rows_r.append(ft.Row(controls=header_r, spacing=2))

        for i, t in enumerate(self.tasks):
            row_ctrls = [
                self.create_bulk_action_cell('row', 'R', t, width=CELL_W_ACTION, height=CELL_H),
                cell_name(t, width=CELL_W_TASK_LABEL)
            ]
            for j, h in enumerate(self.indices_horas):
                val = self._get_val_from_memory_or_json('R', t, h)
                # Create the Excel-style Input
                row_ctrls.append(self.create_excel_input(t, h, val, i, j, width=CELL_W_R, height=CELL_H, font_size=FONT_SIZE))
            rows_r.append(ft.Row(controls=row_ctrls, spacing=2))

        # 3. SKILLS/QUALIFICATIONS (Q)
        rows_q = []
        header_q_act = [ft.Container(width=CELL_W_ACTION + CELL_W_NAME, height=CELL_H)]
        for t in self.tasks:
            header_q_act.append(self.create_bulk_action_cell('col', 'Q', t, width=CELL_W_TASK, height=CELL_H))
        rows_q.append(ft.Row(controls=header_q_act, spacing=2))

        header_q = [ft.Container(width=CELL_W_ACTION, height=CELL_H), cell_header("Person", CELL_W_NAME)]
        for t in self.tasks:
            header_q.append(cell_header(t, CELL_W_TASK))
        rows_q.append(ft.Row(controls=header_q, spacing=2))

        for pers in self.people:
            row_ctrls = [
                self.create_bulk_action_cell('row', 'Q', pers, width=CELL_W_ACTION, height=CELL_H),
                cell_name(pers)
            ]
            for t in self.tasks:
                row_ctrls.append(self.create_cell_button_scaled("YES", "NO", 'Q', pers, t, width=CELL_W_TASK, height=CELL_H, font_size=FONT_SIZE))
            rows_q.append(ft.Row(controls=row_ctrls, spacing=2))

        # 4. MANDATORY/FIXED (F)
        list_f = []
        for t in self.tasks:
            color_task = self.task_colors.get(t, self.COLOR_HEADER_BG)
            
            list_f.append(
                ft.Container(
                    content=ft.Text(f" {t} ", size=12, color="black", weight="bold"),
                    bgcolor=color_task,
                    padding=2,
                    border_radius=3
                )
            )
            
            task_rows = []
            h_row = [cell_header("Person", CELL_W_NAME)]
            for h in self.indices_horas:
                h_row.append(cell_header(f"{self.possible_hours[h]:02d}h", CELL_W_HOUR))
            task_rows.append(ft.Row(controls=h_row, spacing=2))

            for pers in self.people:
                r_row = [cell_name(pers)]
                for h in self.indices_horas:
                    r_row.append(self.create_cell_button_scaled("YES", "", 'F', pers, t, h, width=CELL_W_HOUR, height=CELL_H, font_size=FONT_SIZE))
                task_rows.append(ft.Row(controls=r_row, spacing=2))
            
            list_f.append(ft.Column(task_rows, spacing=2))
            list_f.append(ft.Divider(height=20, color="transparent"))

        container_d.controls = rows_d
        container_r.controls = rows_r
        container_q.controls = rows_q
        container_f.controls = list_f

        def title_separator(text):
            return ft.Container(
                content=ft.Text(text, size=14, weight="bold", color=self.COLOR_TEXT_HIGHLIGHT),
                padding=ft.padding.only(top=20, bottom=5)
            )

        self.content_matrices.controls = [
            title_separator("1. Availability (D)"), 
            ft.Row([container_d], scroll=ft.ScrollMode.AUTO),
            title_separator("2. Requirements (R) - Editable Grid (Tab/Enter)"), 
            ft.Row([container_r], scroll=ft.ScrollMode.AUTO),
            title_separator("3. Skills/Qualifications (Q)"), 
            ft.Row([container_q], scroll=ft.ScrollMode.AUTO),
            ft.Container(height=30)
        ]

        self.content_mandatory.controls = [
            ft.Row([container_f], scroll=ft.ScrollMode.AUTO),
        ]

        self.status_text.value = "Matrices generated."
        if self.page: self.page.update()

    def create_cell_button_scaled(self, label_active, label_inactive, tipo, k1, k2, k3=None, width=60, height=28, font_size=10):
        val = self._get_val_from_memory_or_json(tipo, k1, k2, k3)
        if tipo == 'D': self.state_D.setdefault(k1, {})[k2] = val
        elif tipo == 'Q': self.state_Q.setdefault(k1, {})[k2] = val
        elif tipo == 'F': self.state_F.setdefault(k1, {}).setdefault(k2, {})[k3] = val

        bg_color = "white"
        text_color = "grey"
        text_value = label_inactive

        if val:
            if tipo == 'F':
                bg_color = self.task_colors.get(k2, self.COLOR_ACTIVE)
                text_color = "black" 
                text_value = ""
            else:
                bg_color = self.COLOR_ACTIVE
                text_color = self.TEXT_ACTIVE
                text_value = label_active
        else:
            if tipo == 'D' or tipo == 'Q':
                bg_color = self.COLOR_INACTIVE
                text_color = self.TEXT_INACTIVE
                text_value = label_inactive
            elif tipo == 'F':
                bg_color = "white"
                text_value = ""

        container = ft.Container(
            width=width, height=height, bgcolor=bg_color, border=None, border_radius=3,
            alignment=ft.alignment.center,
            content=ft.Text(text_value, size=font_size, color=text_color),
            data={'tipo': tipo, 'k1': k1, 'k2': k2, 'k3': k3},
            on_click=self.toggle_matrix_btn
        )
        
        if tipo == 'D': self.grid_controls['D'][(k1, k2)] = container
        elif tipo == 'Q': self.grid_controls['Q'][(k1, k2)] = container
            
        return container

    def create_excel_input(self, t, h, val_initial, row_idx, col_idx, width=70, height=20, font_size=10):
        """
        Creates an Excel-like text input cell.
        - Updates state instantly.
        - Tab moves right (standard Flet).
        - Enter moves down (custom implementation).
        - Allow Empty String visuals but treat as 0 for model.
        - Auto-select content on focus.
        - Restrict input to digits only.
        """
        def on_change(e):
            val_str = e.control.value
            # Si se borra todo, el estado es 0, pero permitimos que visualmente quede vacío
            if not val_str:
                new_val = 0
            else:
                try:
                    new_val = int(val_str)
                except ValueError:
                    new_val = 0
            self.state_R.setdefault(t, {})[h] = new_val

        def on_focus(e):
            # Seleccionar todo el texto al hacer foco para sobrescribir fácilmente
            e.control.selection_start = 0
            e.control.selection_end = len(e.control.value)
            e.control.update()

        def on_submit(e):
            # Mover foco a la fila de abajo en la misma columna
            next_row = row_idx + 1
            next_cell = self.r_cells.get((next_row, col_idx))
            if next_cell:
                next_cell.focus()

        # === CAMBIO: Si el valor inicial es 0, mostramos "" ===
        display_val = str(val_initial) if val_initial != 0 else ""

        txt_field = ft.TextField(
            value=display_val,
            text_size=font_size,
            width=width,
            height=height,
            content_padding=ft.padding.only(bottom=21), # Padding ajustado
            text_align=ft.TextAlign.CENTER,
            border=ft.InputBorder.NONE,
            keyboard_type=ft.KeyboardType.NUMBER,
            # Filtro estricto: solo permite dígitos del 0 al 9
            input_filter=ft.InputFilter(allow=True, regex_string=r"^[0-9]*$", replacement_string=""),
            on_change=on_change,
            on_focus=on_focus, # Trigger selección automática
            on_submit=on_submit
        )
        
        # Guardar referencia para navegación
        self.r_cells[(row_idx, col_idx)] = txt_field

        return ft.Container(
            content=txt_field,
            width=width,
            height=height,
            bgcolor="white",
            border=ft.border.all(1, "#e0e0e0"),
            border_radius=5 # Bordes redondeados
        )

    def run_optimization_thread(self, e):
        if not self.indices_horas:
            self.status_text.value = "Error: No hours."
            self.status_text.update()
            return

        self.btn_optimize.disabled = True
        self.progress_bar.visible = True 
        self.status_text.value = "Optimizing (this may take a while)..."
        self.page.update()

        t = threading.Thread(target=self._run_optimization)
        t.start()

    def _run_optimization(self):
        try:
            model, X, W, W_max, W_min = self.gather_data_and_solve()
            status_txt = LpStatus[model.status]
            self.status_text.value = f"Finished: {status_txt}"
            self.show_results_dialog(model, X, W, W_max, W_min)
        except Exception as ex:
            import traceback
            traceback.print_exc()
            self.status_text.value = f"Error: {str(ex)}"
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Critical Error: {str(ex)}"), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
        finally:
            self.btn_optimize.disabled = False
            self.progress_bar.visible = False 
            self.page.update()

    def gather_data_and_solve(self):
        def get_val(ctrl):
            try: return float(ctrl.controls[1].content.value)
            except: return 0.0

        # R data is already in self.state_R updated via on_change
        # Ensure structure, treating Missing/None as 0 explicitly
        R_data = {}
        for t in self.tasks:
            R_data[t] = {}
            for h in self.indices_horas:
                val = self.state_R.get(t, {}).get(h, 0)
                R_data[t][h] = val

        F_save = {}
        for pers in self.people:
            F_save[pers] = {}
            for t in self.tasks:
                F_save[pers][t] = {}
                for h in self.indices_horas:
                    F_save[pers][t][str(h)] = self.state_F.setdefault(pers, {}).setdefault(t, {}).get(h, 0)

        final_data = {
            'people': self.people, 'tasks': self.tasks, 'hours': self.indices_horas,
            'D': {k: {str(idx): v for idx, v in d.items()} for k, d in self.state_D.items()},
            'Q': self.state_Q,
            'R': {k: {str(idx): v for idx, v in d.items()} for k, d in R_data.items()},
            'F': F_save,
            'alpha': get_val(self.in_alpha), 'beta': get_val(self.in_beta), 'gamma': get_val(self.in_gamma),
            'epsilon': get_val(self.in_epsilon), 'timelimit': int(get_val(self.in_timelimit))
        }
        save_data(final_data)

        solver_data = final_data.copy()
        solver_data['D'] = {k: {int(idx): v for idx, v in d.items()} for k, d in final_data['D'].items()}
        solver_data['R'] = {k: {int(idx): v for idx, v in d.items()} for k, d in final_data['R'].items()}
        F_solver = {}
        for i in self.people:
            F_solver[i] = {}
            for t in self.tasks:
                F_solver[i][t] = {}
                for h in self.indices_horas: 
                    F_solver[i][t][h] = self.state_F.get(i, {}).get(t, {}).get(h, 0)
        solver_data['F'] = F_solver

        return solve_model(solver_data)

    def save_excel_results(self, X, W):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Staffing Plan"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            border_style = Side(border_style="thin", color="000000")
            full_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            headers = ["Person"] + [f"{self.possible_hours[h]:02d}h" for h in self.indices_horas] + ["Total"]
            ws.append(headers)

            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = full_border
                ws.column_dimensions[get_column_letter(col_num)].width = 15

            for i in self.people:
                row_idx = ws.max_row + 1
                
                cell_name = ws.cell(row=row_idx, column=1, value=i)
                cell_name.font = Font(bold=True)
                cell_name.border = full_border

                for idx_h, h in enumerate(self.indices_horas):
                    assigned_task = ""
                    color_hex = "FFFFFF"
                    
                    is_available = self._get_val_from_memory_or_json('D', i, h)
                    if not is_available:
                        color_hex = self.FLET_TO_HEX["red100"]

                    for t in self.tasks:
                        if value(X[i][t][h]) == 1:
                            assigned_task = t
                            flet_color = self.task_colors.get(t, "white")
                            color_hex = self.FLET_TO_HEX.get(flet_color, "FFFFFF")
                            break
                    
                    cell = ws.cell(row=row_idx, column=idx_h + 2, value=assigned_task)
                    cell.alignment = center_align
                    cell.border = full_border
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

                val_load = value(W[i]) if value(W[i]) is not None else 0
                cell_total = ws.cell(row=row_idx, column=len(self.indices_horas) + 2, value=int(val_load))
                cell_total.font = Font(bold=True)
                cell_total.alignment = center_align
                cell_total.border = full_border

            filename = "staffing_plan.xlsx"
            wb.save(filename)
            
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Saved: {filename}"), bgcolor="green")
            self.page.snack_bar.open = True
            self.page.update()

            try:
                if platform.system() == 'Darwin':
                    subprocess.call(('open', filename))
                elif platform.system() == 'Windows':
                    os.startfile(filename)
                else:
                    subprocess.call(('xdg-open', filename))
            except Exception as e_open:
                print(f"Could not auto-open file: {e_open}")

        except Exception as ex:
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Excel Error: {str(ex)}"), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()

    def show_results_dialog(self, model, X, W, W_max, W_min):
        status_txt = LpStatus[model.status]
        
        if status_txt != "Optimal":
            content_dlg = ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.WARNING, color="red", size=40),
                    ft.Text(f"No feasible solution found.\nStatus: {status_txt}", size=16, color="red", weight="bold", text_align="center"),
                    ft.Text("Please check constraints (Availability and Requirements) and try again.", text_align="center")
                ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                height=200, alignment=ft.alignment.center
            )
            title_dlg = ft.Text("Attention", color="red")
            actions_dlg = [ft.TextButton("Close", on_click=lambda e: self.page.close(dlg))]
            
            dlg = ft.AlertDialog(
                title=title_dlg, 
                content=content_dlg,
                actions=actions_dlg,
                shape=ft.RoundedRectangleBorder(radius=5),
                modal=True,
            )
            self.page.open(dlg)
            self.page.update()
            return
        
        # === CASO ÓPTIMO ===
        val_wmax = value(W_max) if value(W_max) is not None else 0
        val_wmin = value(W_min) if value(W_min) is not None else 0

        zoom_state = {"scale": 1.0}
        
        BASE_W_NAME = 100
        BASE_W_HOUR = 50
        BASE_W_TOTAL = 50
        BASE_H_ROW = 20
        BASE_FONT_SIZE = 11
        BASE_FONT_SIZE_SMALL = 10
        
        num_people = len(self.people)
        num_hours = len(self.indices_horas)
        
        content_width = BASE_W_NAME + (num_hours * (BASE_W_HOUR + 2)) + BASE_W_TOTAL + 40
        content_height = (num_people + 1) * (BASE_H_ROW + 2) + 40
        
        MIN_WIDTH = 400
        MAX_WIDTH = 1200
        MIN_HEIGHT = 150
        MAX_HEIGHT = 600
        
        dialog_width = max(MIN_WIDTH, min(MAX_WIDTH, content_width))
        dialog_height = max(MIN_HEIGHT, min(MAX_HEIGHT, content_height))

        table_container = ft.Column(spacing=2)
        zoom_label = ft.Text(f"100%", size=12, weight="bold", width=50, text_align=ft.TextAlign.CENTER)

        def build_table(scale):
            W_NAME = int(BASE_W_NAME * scale)
            W_HOUR = int(BASE_W_HOUR * scale)
            W_TOTAL = int(BASE_W_TOTAL * scale)
            H_ROW = int(BASE_H_ROW * scale)
            FONT_SIZE = int(BASE_FONT_SIZE * scale)
            FONT_SIZE_SMALL = int(BASE_FONT_SIZE_SMALL * scale)

            def make_res_cell(content, width, bgcolor="white"):
                return ft.Container(
                    content=content, width=width, height=H_ROW, bgcolor=bgcolor,
                    alignment=ft.alignment.center, border=None, border_radius=3
                )

            rows = []
            
            header_cells = [make_res_cell(ft.Text("Person", weight="bold", size=FONT_SIZE), W_NAME, bgcolor="#F2F2F2")]
            for h in self.indices_horas:
                header_cells.append(make_res_cell(ft.Text(f"{self.possible_hours[h]:02d}h", weight="bold", size=FONT_SIZE), W_HOUR, bgcolor="#F2F2F2"))
            header_cells.append(make_res_cell(ft.Text("Total", weight="bold", size=FONT_SIZE), W_TOTAL, bgcolor="#F2F2F2"))
            rows.append(ft.Row(header_cells, spacing=2))

            for i in self.people:
                row_cells = []
                row_cells.append(ft.Container(
                    content=ft.Text(i, size=FONT_SIZE, weight="bold"), width=W_NAME, height=H_ROW, bgcolor="white",
                    alignment=ft.alignment.center_left, padding=ft.padding.only(left=5), border=None, border_radius=3
                ))
                for idx in self.indices_horas:
                    assigned_task = ''
                    bg_color = "white"
                    
                    is_available = self._get_val_from_memory_or_json('D', i, idx)
                    if not is_available:
                        bg_color = "red100"

                    for t in self.tasks:
                        if value(X[i][t][idx]) == 1:
                            assigned_task = t
                            bg_color = self.task_colors[t]
                            break
                    
                    row_cells.append(make_res_cell(ft.Text(assigned_task, color="black", size=FONT_SIZE_SMALL, weight="bold"), W_HOUR, bgcolor=bg_color))
                
                val_load = value(W[i]) if value(W[i]) is not None else 0
                row_cells.append(make_res_cell(ft.Text(str(int(val_load)), weight="bold", size=FONT_SIZE), W_TOTAL))
                rows.append(ft.Row(row_cells, spacing=2))

            return rows

        def update_table():
            table_container.controls = build_table(zoom_state["scale"])
            zoom_label.value = f"{int(zoom_state['scale'] * 100)}%"
            table_container.update()
            zoom_label.update()

        def zoom_in(e):
            if zoom_state["scale"] < 2.0:
                zoom_state["scale"] = min(2.0, zoom_state["scale"] + 0.1)
                update_table()

        def zoom_out(e):
            if zoom_state["scale"] > 0.5:
                zoom_state["scale"] = max(0.5, zoom_state["scale"] - 0.1)
                update_table()

        def zoom_reset(e):
            zoom_state["scale"] = 1.0
            update_table()

        table_container.controls = build_table(zoom_state["scale"])

        zoom_bar = ft.Row(
            controls=[
                ft.IconButton(icon=ft.Icons.ZOOM_OUT, icon_size=20, tooltip="Zoom Out", on_click=zoom_out),
                zoom_label,
                ft.IconButton(icon=ft.Icons.ZOOM_IN, icon_size=20, tooltip="Zoom In", on_click=zoom_in),
                ft.IconButton(icon=ft.Icons.REFRESH, icon_size=18, tooltip="Reset", on_click=zoom_reset),
            ],
            spacing=0,
        )

        content_dlg = ft.Container(
            content=ft.Row(
                controls=[
                    ft.Column(
                        controls=[table_container],
                        scroll=ft.ScrollMode.AUTO,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        expand=True,
                    )
                ],
                scroll=ft.ScrollMode.ALWAYS,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                alignment=ft.MainAxisAlignment.CENTER,
                expand=True,
            ),
            width=dialog_width,
            height=dialog_height,
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
            alignment=ft.alignment.center,
        )
        
        title_dlg = ft.Text(f"Optimal Plan | Load Max-Min: {int(val_wmax - val_wmin)}", size=16, weight="bold")
        
        actions_dlg = [
            ft.Row(
                controls=[
                    zoom_bar,
                    ft.Container(expand=True),
                    ft.ElevatedButton("Download Excel", icon=ft.Icons.DOWNLOAD, 
                                      on_click=lambda e: self.save_excel_results(X, W)),
                    ft.TextButton("Close", on_click=lambda e: self.page.close(dlg))
                ],
                alignment=ft.MainAxisAlignment.START,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                expand=True,
            )
        ]

        dlg = ft.AlertDialog(
            title=title_dlg, 
            content=content_dlg,
            actions=actions_dlg,
            actions_alignment=ft.MainAxisAlignment.CENTER,
            shape=ft.RoundedRectangleBorder(radius=5),
            modal=True,
        )
        self.page.open(dlg)
        self.page.update()
    

if __name__ == "__main__":
    app = StaffingApp()
    ft.app(target=app.main, view=ft.AppView.FLET_APP)